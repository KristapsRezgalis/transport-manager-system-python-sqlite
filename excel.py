from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import get_column_letter
import datetime
import copy
from pdf import get_forwarder_sender_delivery_data
from my_geopy import get_km


def parse_sap_po(value):
    """
    Tries to store SAP PO as a real number if it's a single value.
    Falls back to text if it contains multiple numbers (e.g. '264889; 264890').
    """
    if value is None:
        return ""

    value_str = str(value).strip()

    if not value_str:
        return ""

    # If it contains a separator, it's multiple POs — keep as text
    if ';' in value_str or ',' in value_str:
        return value_str

    # Try converting to an integer
    try:
        return int(value_str)
    except ValueError:
        # Not a clean number (e.g. contains letters) — keep as text
        return value_str
    
def find_row_by_agreement(ws, table, nr):
    """
    Searches the table's 'Agreement' column for a row matching `nr`.
    Returns the row number if found, otherwise None.
    """
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    header_cells = next(ws.iter_rows(min_row=min_row, max_row=min_row,
                                       min_col=min_col, max_col=max_col))
    col_map = {cell.value: cell.column for cell in header_cells}
    agreement_col = col_map['Agreement']

    # Compare as strings so "123" and 123 both match consistently
    target = str(nr).strip()

    for row_num in range(min_row + 1, max_row + 1):
        cell_value = ws.cell(row=row_num, column=agreement_col).value
        if cell_value is not None and str(cell_value).strip() == target:
            return row_num

    return None

def add_or_update_row_in_table(nr, data):
    filepath = r"C:\Users\kristaps.rezgalis\Gemoss SIA\Transports - Dokumenti\Transports 2025.xlsx"
    sheet_name = "Sep 2026 "
    table_name = "TabulaSep2026"

    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    table = ws.tables[table_name]

    df_fw, df_fw_contact, df_sender_company, df_sender_company_address, df_sender_company_contact, df_delivery_company, df_delivery_company_address, df_delivery_company_contact = get_forwarder_sender_delivery_data(data)
    address_load = f"{df_sender_company_address['adr_city'].iloc[0]}, {df_sender_company_address['adr_post_code'].iloc[0]}"
    address_deliver = f"{df_delivery_company_address['adr_city'].iloc[0]}, {df_delivery_company_address['adr_post_code'].iloc[0]}"

    row_dict = {
        "Nosūtītājs": data.get('sender'),
        "Adrese": f"{df_sender_company_address['adr_street'].iloc[0]}, {df_sender_company_address['adr_city'].iloc[0]}, {df_sender_company_address['adr_post_code'].iloc[0]}",
        "Valsts": df_sender_company_address['adr_country'].iloc[0],
        "Piegāde": df_delivery_company_address['adr_name'].iloc[0],
        "Paletes": data.get('pallets'),
        "Svars": data.get('weight'),
        "Izmērs": "",
        "LDM": data.get('ldm'),
        "Iekraušana līdz": datetime.datetime.strptime(data.get('loading_to'), '%Y-%m-%d'),
        "Izkraušana līdz": datetime.datetime.strptime(data.get('unloading_to'), '%Y-%m-%d'),
        "Pārvadātājs": data.get('forwarder'),
        "Transporta cena, EUR": data.get('cost'),
        "SAP PO": parse_sap_po(data.get('sap_po')),
        "Iepircējs": data.get('purch_manager'),
        "Agreement": nr,
        "Documents": data.get('doc_loc'),
        "km": get_km(address_load, address_deliver),
        "REFS": f"{data.get('temp_min')}...{data.get('temp_max')}°C" if data.get('temp_min') else ""
    }

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    header_cells = next(ws.iter_rows(min_row=min_row, max_row=min_row,
                                       min_col=min_col, max_col=max_col))
    col_map = {cell.value: cell.column for cell in header_cells}

    for key in row_dict:
        if key not in col_map:
            raise ValueError(f"'{key}' is not a real column name. Available: {list(col_map.keys())}")

    def copy_format(source_row, col_num, target_cell):
        template_cell = ws.cell(row=source_row, column=col_num)
        target_cell.number_format = template_cell.number_format
        target_cell.font = copy.copy(template_cell.font)
        target_cell.border = copy.copy(template_cell.border)
        target_cell.fill = copy.copy(template_cell.fill)
        target_cell.alignment = copy.copy(template_cell.alignment)

    existing_row = find_row_by_agreement(ws, table, nr)

    if existing_row is not None:
        # --- UPDATE existing row ---
        target_row_number = existing_row
        for column_name, value in row_dict.items():
            col_num = col_map[column_name]
            ws.cell(row=target_row_number, column=col_num, value=value)
        # No need to touch formatting or table.ref — the row already exists and is already formatted
        print(f"Updated existing row {target_row_number} (Agreement={nr})")

    else:
        # --- INSERT new row (your original logic) ---
        target_row_number = max_row + 1
        template_row_number = max_row

        for col_num in range(min_col, max_col + 1):
            cell = ws.cell(row=target_row_number, column=col_num)
            copy_format(template_row_number, col_num, cell)

        for column_name, value in row_dict.items():
            col_num = col_map[column_name]
            ws.cell(row=target_row_number, column=col_num, value=value)

        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{target_row_number}"
        print(f"Added new row {target_row_number} (Agreement={nr})")

    # This runs either way — the CO2 formula always needs to reference the correct row
    svars_col_letter = get_column_letter(col_map['Svars'])
    km_col_letter = get_column_letter(col_map['km'])
    ws.cell(row=target_row_number, column=col_map['CO2 kg'],
            value=f"=({svars_col_letter}{target_row_number}/1000)*{km_col_letter}{target_row_number}*0.1")

    wb.save(filepath)
